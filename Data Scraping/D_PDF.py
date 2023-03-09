import sys
import os

os.chdir('/Users/kbiscoch/Documents/Research_remote/GitHub/HP_Modelling/Data Scraping')

##START - FIRST Page of FIRST HP model u want to study
##END - FIRST Page of LAST HP model u want to study
##HEAT - Heating values page
START = 4
END = 13
HEAT = 12

try:
    import CoolProp.CoolProp as CP #package that gets thermodynamic properties for stuff
    import tabula as tb
    import pandas as pd
    import re
    import numpy as np
    from PyPDF2 import PdfFileWriter, PdfFileReader
    import openpyxl
    import matplotlib.pyplot as plt
    from sklearn import datasets, linear_model
    import pyromat as pm
    from sklearn.metrics import mean_squared_error, r2_score
    import os
except:
    sys.exit("MISSING NEEDED PACKAGES")

##Plots all given info
##Parameters:
##c - slopes of cooling
##i - intrcpts of cooling
##ch - slopes of Heating
##ih - intrcpts of heating
def plot_results(c,i,ch,ih):
    x = np.linspace(60,120)
    x_h = np.linspace(-10,65)
    
    plt.figure(1)
    
    for cc,ic,ch,ih in zip(c, i, ch, ih):
        v = (x * cc + ic)
        v_h = (x_h * ch + ih)

        plt.plot(x,v)
        plt.plot(x_h,v_h)
    
    plt.title("COPs vs OD Temperature (DF System)")
    plt.xlabel("Temperature (F)")
    plt.ylabel("COP")
    
##Prints all given info
##Parameters:
##coeff_c - slopes of cooling
##intrcpt_c - intrcpts of cooling
##coeff_h - slopes of Heating
##intrcpt_h - intrcpts of heating
def print_results(coeff_c, intrcpt_c, coeff_h, intrcpt_h):
    for cc,ic,ch,ih in zip(coeff_c, intrcpt_c, coeff_h, intrcpt_h):
        print("COP = T * " + str(cc) + " + " + str(ic))
        print("COP = T * " + str(ch) + " + " + str(ih))
        print("\n")
    
#Makes the excel file to use the data
#Parameter:
#file - data file (pdf)
def generateExcel(file):
    
    data = tb.read_pdf(file,area = [50,0,555,780], pages=str(START)+"-"+str(END),lattice=False,pandas_options={'header': None},stream=True)
    heating_data = tb.read_pdf(file, pages=HEAT,pandas_options={'header': None},stream=True) # pages argument may change with every PDF
    
    #Cooling data
    i = 0
    with pd.ExcelWriter('output.xlsx') as writer:   
        for dataset in data:
            dataset = dataset.replace(np.nan, '',regex=True)
            data[i] = dataset
            i=i+1
            dataset.to_excel(writer, sheet_name='Sheet_name'+str(i+1))
    i=0

    #Heating Data
    with pd.ExcelWriter('output_h.xlsx') as writer:   
        for dataset in heating_data:
            dataset = dataset.replace(np.nan, '',regex=True)
            heating_data[i] = dataset
            i=i+1
            dataset.to_excel(writer, sheet_name='Sheet_name'+str(i+1))
            
    wrbk = openpyxl.load_workbook("output.xlsx")
    wrbk_heating = openpyxl.load_workbook("output_h.xlsx")
    
    return wrbk, wrbk_heating


##Paramaters are return values of top fuction
#This function basically gets the important info from the data

def stripData(wrbk, wrbk_heating):
    kw = []
    MBh = []
    kws = []
    MBhs = []
    k = 0
    for sh in wrbk:
        i = 0
        for row in sh.iter_rows():
            if(i == 0):
                if(len(MBh) != 0):
                    MBhs.append(MBh)
                if(len(kw) != 0):
                    kws.append(kw)
                kw = []
                MBh = []
            k = 0
            for cell in row:
                if k == 1:
                    if(type(cell.value)==int or type(cell.value) == float):
                        MBh.append((cell.value))
                    else:
                        try:
                            MBh.extend(list(map(float, cell.value.split())))                 
                        except:
                            if cell.value != None:
                                x = cell.value
                                x = x.replace('-','')
                                MBh.extend(list(map(float, x.split())))   
                            pass

                if k == 2:
                    # print(cell.value)
                    if(type(cell.value)==int or type(cell.value) == float):
                        kw.append((cell.value))
                    else:
                        try:
                            kw.extend(list(map(float, cell.value.split())))
                        except:
                            if cell.value != None:
                                x = cell.value
                                x = x.replace('-','')
                                kw.extend(list(map(float, x.split())))   
                            pass

                if cell.value == "MBh":
                    k = 1
                if (type(cell.value) != None) and "kW" in str(cell.value):
                    k = 2
            i = i + 1
            if i == 6:
                i = 0

    kws.append(kw)
    
    return kws, MBhs

        

##Gets COPS of heating section, they are explicitly stated so i just take them
def getHeatingCOPs(wrbk_heating):

    COPs_heating = []

    for sh in wrbk_heating:
        COP = []
        for row in sh.iter_rows():
            k = -1
            for cell in row:
                if k == 9:
                    COP.extend(map(float,str(cell.value).split()))
                if cell.value == "COP":
                    k = 9
        COPs_heating.append(COP)
    
    return COPs_heating

#Gets COPs from MBhs (Heat) and kws (Energy)
def getCOPs(MBhs,kws,con_rate):
    COPs = []
    for i in range(len(MBhs)):
        COPs.append((np.array(MBhs[i]) * con_rate) / np.array(kws[i]))
        
    return COPs

#Does linear regression on the COPs and returns the slopes and intercepts of both cooling and heating
def getCoeffs(COPs,COPs_heating,temps,heating_temps):
    i = 0
    x = np.linspace(65,115)[:,None]
    x_h = np.linspace(-10,65)[:,None]

    coeffs = []
    intrcpts = []
    coeffs_h = []
    intrcpts_h = []
    regr = linear_model.LinearRegression()
    
    
    for copset in COPs[::12]:  

        regr.fit(np.array(temps)[:,None],copset[2::3])
        

        coeffs.append(regr.coef_)
        intrcpts.append(regr.intercept_)
        y = x * regr.coef_ + regr.intercept_
        y_r = (temps * regr.coef_ + regr.intercept_)
        r2 = r2_score(copset[2::3],y_r)
        print("Cooling r^2 = " + str(r2))
        
        
        # print(np.array(heating_temps[::-1])[:,None])
        # print(COPs_heating[i][::-1])
        regr.fit(np.array(heating_temps[::-1])[:,None], COPs_heating[i][::-1])
        y_hr = (heating_temps[::-1] * regr.coef_ + regr.intercept_)
        r2 = r2_score(COPs_heating[i][::-1],y_hr)

        coeffs_h.append(regr.coef_)
        intrcpts_h.append(regr.intercept_)
        print("Heating r^2 = " + str(r2))
        
        print("\n")
        
        i = i + 1
        
    return coeffs, intrcpts, coeffs_h, intrcpts_h

def main():
    ##Initialize Variables
    kws = []
    MBhs = []
    kw = []
    MBh = []
    kws_heating = []
    MBhs_heating = []
    kw_heating = []
    MBh_heating = []
    slopes = []
    y_intrcpt = []
    slopes_heating = []
    y_intrcpts_heating = []
    sat_dep_slps = []
    r2_scores = []
    plotting=True
    
    ##Conversion rate of MBh --> kW
    con_rate = 0.2930710702

    ##Temperatures that loaded cooling data is on -- Could also use renaming
    cooling_temps = [65,75,85,95,105,115]

    ##Temperatures that loaded heating data is on
    heating_temps = [65,60,55,50,47,45,40,35,30,25,20,17,15,10,5,0,-5,-10]

    file = input("Enter file name with PDF extenstion included: ") # R410A_DZ20VC.pdf
    wrbk, wrbk_heating = generateExcel(file)
    kws, MBhs = stripData(wrbk,wrbk_heating)
    COPs = getCOPs(MBhs, kws, con_rate)
    COPs_heating = getHeatingCOPs(wrbk_heating) # returns empty list for DZ20VC
    
    c,i,ch,ih = getCoeffs(COPs,COPs_heating,cooling_temps,heating_temps)
    
##For average heating and cooling line at T_ref = 65, uncomment this
#    a,b,x,y = 0,0,0,0
#    for k in range(len(c)):
#        a = a + c[k]
#        b = b + 65 * c[k] + i[k]
#        x = x + ch[k]
#        y = y + 65 * ch[k] + ih[k]
#
#    a = a / len(c)
#    b = b / len(c)
#    x = x / len(c)
#    y = y / len(c)
#
#    print(a,b,x,y)
    # print_results()
    print_results(c, i, ch, ih)
    if plotting == True:
        plot_results(c,i,ch,ih)

    
    # print(c)
    
if __name__ == '__main__':
    main()
